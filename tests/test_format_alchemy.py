import os
import sqlite3
from pathlib import Path
from unittest.mock import patch, MagicMock

import pytest

try:
    import openpyxl
    has_openpyxl = True
except ImportError:
    has_openpyxl = False

from data_fetcher.format_alchemy import FormatAlchemyEngine


@pytest.fixture
def temp_csv_path(tmp_path: Path) -> str:
    """Fixture to create a temporary basic CSV file."""
    import pandas as pd
    csv_file = tmp_path / "test_data_raw.csv"
    df = pd.DataFrame({"col1": [1, 2, 3], "col2": ["a", "b", "c"]})
    df.to_csv(csv_file, index=False)
    return str(csv_file)


@pytest.fixture
def null_csv_path(tmp_path: Path) -> str:
    """Fixture to create a temporary CSV file with an entirely null column."""
    import pandas as pd
    csv_file = tmp_path / "null_data.csv"
    df = pd.DataFrame({"col1": [1, 2, 3], "col2": [None, None, None]})
    df.to_csv(csv_file, index=False)
    return str(csv_file)


@pytest.fixture
def long_name_csv_path(tmp_path: Path) -> str:
    """Fixture to create a temporary CSV file with a name exceeding 31 characters."""
    import pandas as pd
    csv_file = tmp_path / ("a" * 40 + ".csv")
    df = pd.DataFrame({"col1": [1]})
    df.to_csv(csv_file, index=False)
    return str(csv_file)


def test_csv_to_sqlite_basic(temp_csv_path: str) -> None:
    """Test standard ingestion of a CSV into a SQLite database."""
    import pandas as pd
    engine = FormatAlchemyEngine(temp_csv_path)
    engine.csv_to_sqlite()

    assert os.path.exists(engine.db_filepath)
    with sqlite3.connect(engine.db_filepath) as conn:
        df_db = pd.read_sql_query(f'SELECT * FROM "{engine.dataset_name}"', conn)

    assert len(df_db) == 3
    assert len(df_db.columns) == 2
    assert "col1" in df_db.columns


@pytest.mark.skipif(not has_openpyxl, reason="openpyxl is not installed")
def test_sqlite_to_excel_basic(temp_csv_path: str) -> None:
    """Test full pipeline execution producing a valid Excel file."""
    import pandas as pd
    engine = FormatAlchemyEngine(temp_csv_path)
    engine.execute_pipeline()

    assert os.path.exists(engine.excel_filepath)

    df_excel = pd.read_excel(engine.excel_filepath, engine="openpyxl")
    assert len(df_excel) == 3
    assert len(df_excel.columns) == 2


@pytest.mark.skipif(not has_openpyxl, reason="openpyxl is not installed")
def test_excel_row_cap(tmp_path: Path) -> None:
    """Test Excel export handles row counts exceeding the limit gracefully."""
    import pandas as pd
    temp_csv_path = str(tmp_path / "test_data_raw.csv")
    df = pd.DataFrame({"col1": [1]})
    df.to_csv(temp_csv_path, index=False)
    engine = FormatAlchemyEngine(temp_csv_path)
    engine.csv_to_sqlite()

    mock_cursor = MagicMock()
    mock_cursor.fetchone.return_value = [1048580]
    mock_conn = MagicMock()
    mock_conn.cursor.return_value = mock_cursor
    mock_conn.__enter__.return_value = mock_conn

    with patch("sqlite3.connect", return_value=mock_conn), \
         patch("pandas.read_sql_query", return_value=pd.DataFrame({"col1": [1]})) as mock_read_sql, \
         patch("builtins.print") as mock_print:

        engine.sqlite_to_excel()

        mock_print.assert_any_call(
            "[Warning] Dataset has 1048580 rows, which exceeds Excel's limit. Exporting only the first 1,048,575 rows."
        )
        mock_read_sql.assert_called_once()
        call_args = mock_read_sql.call_args[0]
        assert "LIMIT 1048575" in call_args[0]


@pytest.mark.skipif(not has_openpyxl, reason="openpyxl is not installed")
def test_all_null_column_width(null_csv_path: str) -> None:
    """Test Excel export does not crash when sizing columns with all nulls."""
    import pandas as pd
    engine = FormatAlchemyEngine(null_csv_path)
    engine.execute_pipeline()

    assert os.path.exists(engine.excel_filepath)
    df_excel = pd.read_excel(engine.excel_filepath, engine="openpyxl")
    assert len(df_excel) == 3
    assert df_excel["col2"].isna().all()


@pytest.mark.skipif(not has_openpyxl, reason="openpyxl is not installed")
def test_long_dataset_name_sheet_truncation(long_name_csv_path: str) -> None:
    """Test Excel sheet name truncation to 31 characters."""
    import openpyxl
    engine = FormatAlchemyEngine(long_name_csv_path)
    engine.execute_pipeline()

    assert os.path.exists(engine.excel_filepath)

    wb = openpyxl.load_workbook(engine.excel_filepath)
    assert len(wb.sheetnames) == 1
    sheet_name = wb.sheetnames[0]

    assert len(sheet_name) <= 31
    assert sheet_name == ("a" * 31)


def test_file_not_found() -> None:
    """Test initialization fails when CSV does not exist."""
    non_existent_path = "does_not_exist_file.csv"
    with pytest.raises(FileNotFoundError, match="Source CSV file not found: does_not_exist_file.csv"):
        FormatAlchemyEngine(non_existent_path)
